"""
Excel Parser - Extract financial data from Excel files

This module handles parsing of Excel financial statements (.xlsx, .xls)
with automatic sheet and column detection.

Security Rationale:
- Processes files locally using openpyxl/xlrd
- No external API calls or macros execution
- Input validation for malicious Excel content
- Memory-safe parsing with size limits
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import re
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook

from security.file_handler import SecureFileHandler


class ExcelParser:
    """
    Parse financial data from Excel statements.

    Security: All processing happens locally using openpyxl/pandas.
    Macros are disabled and content is validated for security.
    """

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.secure_handler = SecureFileHandler()

        # Common column patterns (same as CSV parser)
        self.date_columns = [
            "date",
            "transaction_date",
            "trans_date",
            "posted_date",
            "effective_date",
        ]
        self.description_columns = [
            "description",
            "memo",
            "payee",
            "merchant",
            "reference",
            "details",
        ]
        self.amount_columns = [
            "amount",
            "transaction_amount",
            "debit_amount",
            "credit_amount",
        ]
        self.debit_columns = ["debit", "debit_amount", "withdrawal", "payment"]
        self.credit_columns = ["credit", "credit_amount", "deposit", "income"]

        # Date parsing formats
        self.date_formats = [
            "%m/%d/%Y",
            "%m/%d/%y",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%m-%d-%Y",
            "%d-%m-%Y",
        ]

    def parse_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """
        Parse an Excel financial statement.

        Args:
            file_path: Path to Excel file

        Returns:
            List[Dict[str, Any]]: Extracted transaction data

        Security: Uses secure file validation and disables macros.
        """
        try:
            # Validate file first
            validated_path = self.secure_handler.validate_file_path(file_path)
            self.logger.info(f"Parsing Excel file: {validated_path}")

            # Check if file is .xlsx or .xls
            if validated_path.suffix.lower() == ".xlsx":
                transactions = self._parse_xlsx_file(validated_path)
            else:  # .xls
                transactions = self._parse_xls_file(validated_path)

            self.logger.info(f"Parsed {len(transactions)} transactions from Excel")
            return transactions

        except Exception as e:
            self.logger.error(f"Excel parsing failed: {e}")
            return []

    def _parse_xlsx_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """
        Parse .xlsx file using openpyxl.

        Args:
            file_path: Path to .xlsx file

        Returns:
            List[Dict[str, Any]]: Extracted transactions

        Security: Loads workbook with data_only=True to disable formulas.
        """
        transactions = []

        try:
            # Load workbook with security settings
            workbook = load_workbook(
                filename=file_path,
                read_only=True,  # Read-only for security
                data_only=True,  # Don't evaluate formulas
                keep_links=False,  # Don't load external links
            )

            # Find the best worksheet to parse
            worksheet = self._find_transaction_worksheet(workbook)
            if not worksheet:
                self.logger.error("No suitable worksheet found")
                return []

            # Convert worksheet to DataFrame for easier processing
            df = self._worksheet_to_dataframe(worksheet)

            if df is None or df.empty:
                return []

            # Detect column mapping and parse transactions
            column_mapping = self._detect_columns(df)
            if column_mapping:
                transactions = self._parse_transactions(df, column_mapping)

        except Exception as e:
            self.logger.error(f"XLSX parsing failed: {e}")

        return transactions

    def _parse_xls_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """
        Parse .xls file using pandas/xlrd.

        Args:
            file_path: Path to .xls file

        Returns:
            List[Dict[str, Any]]: Extracted transactions
        """
        transactions = []

        try:
            # Read Excel file with pandas (handles .xls files)
            excel_file = pd.ExcelFile(file_path)

            # Try each sheet to find transaction data
            best_sheet = None
            best_score = 0

            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        nrows=1000,  # Limit rows for security
                    )

                    score = self._score_worksheet_for_transactions(df)
                    if score > best_score:
                        best_score = score
                        best_sheet = sheet_name

                except Exception as e:
                    self.logger.debug(f"Error reading sheet {sheet_name}: {e}")
                    continue

            if best_sheet:
                df = pd.read_excel(excel_file, sheet_name=best_sheet, nrows=10000)

                # Detect columns and parse transactions
                column_mapping = self._detect_columns(df)
                if column_mapping:
                    transactions = self._parse_transactions(df, column_mapping)

        except Exception as e:
            self.logger.error(f"XLS parsing failed: {e}")

        return transactions

    def _find_transaction_worksheet(self, workbook) -> Optional[Any]:
        """
        Find the worksheet most likely to contain transaction data.

        Args:
            workbook: openpyxl workbook object

        Returns:
            Optional[Any]: Best worksheet or None
        """
        best_worksheet = None
        best_score = 0

        for worksheet in workbook.worksheets:
            try:
                # Convert to DataFrame for analysis
                df = self._worksheet_to_dataframe(worksheet)

                if df is None or df.empty:
                    continue

                score = self._score_worksheet_for_transactions(df)
                if score > best_score:
                    best_score = score
                    best_worksheet = worksheet

            except Exception as e:
                self.logger.debug(f"Error analyzing worksheet {worksheet.title}: {e}")
                continue

        return best_worksheet

    def _worksheet_to_dataframe(self, worksheet) -> Optional[pd.DataFrame]:
        """
        Convert openpyxl worksheet to pandas DataFrame.

        Args:
            worksheet: openpyxl worksheet

        Returns:
            Optional[pd.DataFrame]: DataFrame or None
        """
        try:
            # Get worksheet data
            data = []
            for row in worksheet.iter_rows(max_row=10000):  # Limit for security
                row_data = []
                for cell in row:
                    value = cell.value
                    # Handle datetime objects
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d")
                    row_data.append(value)
                data.append(row_data)

            if not data:
                return None

            # Create DataFrame with first row as headers
            df = pd.DataFrame(data[1:], columns=data[0])

            # Remove empty rows and columns
            df = df.dropna(how="all").dropna(axis=1, how="all")

            return df

        except Exception as e:
            self.logger.error(f"Failed to convert worksheet to DataFrame: {e}")
            return None

    def _score_worksheet_for_transactions(self, df: pd.DataFrame) -> int:
        """
        Score worksheet for likelihood of containing transaction data.

        Args:
            df: Pandas DataFrame

        Returns:
            int: Score (higher = more likely to contain transactions)
        """
        if df is None or df.empty:
            return 0

        score = 0
        column_names = [str(col).lower() for col in df.columns]

        # Check for financial keywords in column names
        financial_keywords = [
            "date",
            "amount",
            "transaction",
            "debit",
            "credit",
            "description",
            "payee",
            "balance",
            "deposit",
            "withdrawal",
        ]

        for keyword in financial_keywords:
            if any(keyword in col_name for col_name in column_names):
                score += 10

        # Check for date-like columns
        for col in df.columns:
            if self._looks_like_date_column(df[col]):
                score += 15

        # Check for amount-like columns
        for col in df.columns:
            if self._looks_like_amount_column(df[col]):
                score += 10

        # Penalty for too few rows
        if len(df) < 5:
            score -= 20

        return score

    def _detect_columns(self, df: pd.DataFrame) -> Optional[Dict[str, str]]:
        """
        Detect column mapping for financial data.

        Args:
            df: Pandas DataFrame

        Returns:
            Optional[Dict[str, str]]: Column mapping or None
        """
        # Same logic as CSV parser
        columns = [col.lower().strip() for col in df.columns]
        mapping = {}

        # Find date column
        date_col = None
        for col_name in columns:
            if any(date_pattern in col_name for date_pattern in self.date_columns):
                date_col = df.columns[columns.index(col_name)]
                break

        if not date_col:
            for col in df.columns:
                if self._looks_like_date_column(df[col]):
                    date_col = col
                    break

        if not date_col:
            return None

        mapping["date"] = date_col

        # Find description column
        desc_col = None
        for col_name in columns:
            if any(
                desc_pattern in col_name for desc_pattern in self.description_columns
            ):
                desc_col = df.columns[columns.index(col_name)]
                break

        if not desc_col:
            for col in df.columns:
                if col != date_col and df[col].dtype == "object":
                    desc_col = col
                    break

        if not desc_col:
            return None

        mapping["description"] = desc_col

        # Find amount columns
        amount_col = None
        debit_col = None
        credit_col = None

        for col_name in columns:
            if any(
                amount_pattern in col_name for amount_pattern in self.amount_columns
            ):
                amount_col = df.columns[columns.index(col_name)]
                break

        for col_name in columns:
            if any(debit_pattern in col_name for debit_pattern in self.debit_columns):
                debit_col = df.columns[columns.index(col_name)]
            if any(
                credit_pattern in col_name for credit_pattern in self.credit_columns
            ):
                credit_col = df.columns[columns.index(col_name)]

        if amount_col:
            mapping["amount"] = amount_col
        elif debit_col and credit_col:
            mapping["debit"] = debit_col
            mapping["credit"] = credit_col
        else:
            for col in df.columns:
                if col not in [date_col, desc_col] and self._looks_like_amount_column(
                    df[col]
                ):
                    mapping["amount"] = col
                    break

        if "amount" not in mapping and (
            "debit" not in mapping or "credit" not in mapping
        ):
            return None

        self.logger.info(f"Detected Excel column mapping: {mapping}")
        return mapping

    def _looks_like_date_column(self, series: pd.Series) -> bool:
        """Check if series contains date-like data."""
        try:
            sample = series.dropna().head(10)
            if len(sample) == 0:
                return False

            date_count = 0
            for value in sample:
                if self._parse_date_value(value):
                    date_count += 1

            return date_count >= len(sample) * 0.8

        except Exception:
            return False

    def _looks_like_amount_column(self, series: pd.Series) -> bool:
        """Check if series contains amount-like data."""
        try:
            sample = series.dropna().head(10)
            if len(sample) == 0:
                return False

            numeric_count = 0
            for value in sample:
                if self._parse_amount_value(value) is not None:
                    numeric_count += 1

            return numeric_count >= len(sample) * 0.8

        except Exception:
            return False

    def _parse_transactions(
        self, df: pd.DataFrame, column_mapping: Dict[str, str]
    ) -> List[Dict[str, Any]]:
        """Parse transactions from DataFrame."""
        transactions = []

        for index, row in df.iterrows():
            try:
                transaction = self._parse_transaction_row(row, column_mapping)
                if transaction:
                    transactions.append(transaction)
            except Exception as e:
                self.logger.debug(f"Failed to parse Excel row {index}: {e}")
                continue

        return transactions

    def _parse_transaction_row(
        self, row: pd.Series, mapping: Dict[str, str]
    ) -> Optional[Dict[str, Any]]:
        """Parse single transaction row."""
        try:
            # Parse date
            date_value = row[mapping["date"]]
            transaction_date = self._parse_date_value(date_value)
            if not transaction_date:
                return None

            # Parse description
            description = str(row[mapping["description"]]).strip()
            if not description or description.lower() in ["nan", "none", ""]:
                description = "Unknown Transaction"

            # Parse amount
            if "amount" in mapping:
                amount_value = row[mapping["amount"]]
                amount = self._parse_amount_value(amount_value)
                if amount is None:
                    return None

                transaction_type = "debit" if amount < 0 else "credit"
                amount = abs(amount)

            else:
                debit_value = row.get(mapping.get("debit", ""), 0) or 0
                credit_value = row.get(mapping.get("credit", ""), 0) or 0

                debit_amount = self._parse_amount_value(debit_value) or 0
                credit_amount = self._parse_amount_value(credit_value) or 0

                if debit_amount > 0:
                    amount = debit_amount
                    transaction_type = "debit"
                elif credit_amount > 0:
                    amount = credit_amount
                    transaction_type = "credit"
                else:
                    return None

            return {
                "transaction_date": transaction_date,
                "description": description,
                "amount": amount,
                "transaction_type": transaction_type,
                "category_id": None,
                "subcategory": None,
                "notes": None,
                "file_hash": None,
            }

        except Exception as e:
            self.logger.debug(f"Failed to parse Excel transaction row: {e}")
            return None

    def _parse_date_value(self, date_value: Any) -> Optional[datetime]:
        """Parse date value into datetime."""
        if pd.isna(date_value):
            return None

        # Handle datetime objects directly
        if isinstance(date_value, datetime):
            return date_value

        date_str = str(date_value).strip()

        # Try pandas parsing
        try:
            parsed_date = pd.to_datetime(date_str, infer_datetime_format=True)
            return parsed_date.to_pydatetime()
        except Exception:
            pass

        # Try manual format parsing
        for fmt in self.date_formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue

        return None

    def _parse_amount_value(self, amount_value: Any) -> Optional[float]:
        """Parse amount value into float."""
        if pd.isna(amount_value):
            return None

        try:
            # Handle numeric values directly
            if isinstance(amount_value, (int, float)):
                return float(amount_value)

            # Clean string values
            amount_str = str(amount_value).strip()
            clean_amount = re.sub(r"[^\d\.\-\+]", "", amount_str)

            if clean_amount:
                return float(clean_amount)

        except (ValueError, TypeError):
            pass

        return None
